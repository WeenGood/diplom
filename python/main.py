from tkinter import *
from tkinter import filedialog
from tkinter import messagebox

from openpyxl import Workbook
from openpyxl import load_workbook

import re
import sqlite3
import os
from scipy import optimize
import random
#класс главного окна
class Application(Frame):
    #функция инициализации компонентов
    def __init__(self,master):

        super(Application, self).__init__()
        self.master = master
        #создание контейнера для компонентов
        self.main = Frame(master, width = 800, height = 600)
        self.main.pack()
        #заголовок Главного меню
        self.heading = Label(self.main, text="Главное меню")
        self.heading.grid(row = 0, column = 5, pady = 20)
        #кнопка Получить задание
        self.bGetTask = Button(self.main, text = "Получить задание", width = 20, height = 2, command = self.getTaskSummon)        
        self.bGetTask.grid(row = 5, column = 5, pady = 20)
        #кнопка Проверить решение
        self.bSolutionCheck = Button(self.main, text = "Проверить решение", width = 20, height = 2, command = self.wSolutionCheckSummon)        
        self.bSolutionCheck.grid(row = 10, column = 5, pady = 20)
        #кнопка Открыть теорию
        self.bTheory = Button(self.main, text = "Открыть теорию", width = 20, height = 2, command = self.getTheorySummon)        
        self.bTheory.grid(row = 15, column = 5, pady = 20)
        #кнопка Настройки
        self.bSettings = Button(self.main, text = "Настройки", width = 20, height = 2)        
        self.bSettings.grid(row = 20, column = 5, pady = 20)
        #кнопка Выход
        self.bExit = Button(self.main, text = "Выход", width = 20, height = 2, command = self.main.quit)        
        self.bExit.grid(row = 25, column = 5, pady = 20)
    #функция вызова окно проверки рещения
    def wSolutionCheckSummon(self):
        inputDialog = wSolutionCheck(self)
        root.wait_window(inputDialog.top)
    #функция вызова окна получения задания
    def getTaskSummon(self):
        inputDialog = getTask(self)
        root.wait_window(inputDialog.top)

    def getTheorySummon(self):
        inputDialog = getTheory(self)
        root.wait_window(inputDialog.top)

#окно выдачи задания
class getTask:
    def __init__(self,master):
        top = self.top = Toplevel(master)
        main = Frame(top,width = 600, height = 450)
        main.pack()
        lTitle = Label(main, text="Получение задания")
        lTitle.grid(row = 0, column = 5)
        '''
            lBrowse = Label(main, text = "Путь к файлу: ")
            lBrowse.grid(row = 5, column = 4)
            self.ePath = Entry(main)
            self.ePath.grid(row = 5, column = 5)

            bPath = Button(main, text = "Обзор", command = self.browse)
            bPath.grid(row = 5,column = 6)
        '''

        lType = Label(main, text = "Введите номер задания: ")
        lType.grid(row = 1, column = 4)

        self.eType = Entry(main)
        self.eType.grid(row = 1, column = 5)

        bGet = Button(main, text = "Получить задание", command = self.getTask)
        bGet.grid(row = 1,column = 6)

        self.lTaskText = StringVar()
        self.lTask = Label(main, textvariable = self.lTaskText)
        self.lTask.grid(row = 10, column = 4, columnspan = 3)
        
        self.lVariablesText = StringVar()
        self.lVariables = Label(main, textvariable = self.lVariablesText)
        self.lVariables.grid(row = 11, column = 4, columnspan = 3)

        conn = sqlite3.connect('db.db')
        self.c = conn.cursor()

        top.geometry("600x450+0+0")
        top.resizable(0,0)
    '''
    #обзор файла
    def browse(self):
        Tk().withdraw() 
        self.filename = filedialog.askopenfilename()
        self.ePath.delete(0,'end')
        self.ePath.insert(0,self.filename)
    '''
    def getTask(self):
        while 1:
            typeOfTask = self.eType.get()
            if typeOfTask.isdigit() is False:
                messagebox.showinfo('Ошибка!', 'Некорректный номер задания!')
                break
            sql = "SELECT COUNT(*) FROM tasks"
            countTasks = self.c.execute(sql)
            sql = "SELECT id,variables FROM tasks WHERE typeOfTask LIKE " + str(typeOfTask)
            try:
                result = self.c.execute(sql)
            except:
                messagebox.showinfo('Ошибка!', 'Введите номер задания!')
                break

            result = self.transformToList(result)
            
            if len(result) == 0:
                messagebox.showinfo('Ошибка!', 'Некорректный номер задания!')
                break

            newRes = []
            ids = []
            for a in result:
                if result.index(a) %2 != 0:
                    newRes.append(a)
                else:
                    ids.append(a)
            result = newRes

            number = random.randint(0,len(result)-1)
            result = result[number]
            myID = ids[number]
            
            result = result.split(';')
            variables = []
            
            for a in result:
                variables.append(a.split(":"))
            if typeOfTask == '1':
                self.lTaskText.set(''' 
                Производитель может изготавливать продукцию на одном из двух станков.
                Обозначим через x1 количество продукции, изготовленное на станке 1,
                а через x2 количество продукции, изготовленное на станке 2.
                Пусть a1*x1 + b1*x1^2 – стоимость производства на станке 1,
                a2*x2  + b2*x2^2 – стоимость производства на станке 2.
                Необходимо найти значения x1 и x2,
                которые бы минимизировали суммарные затраты при соблюдении требования,
                что суммарное количество произведенной продукции должно быть равно заданному числу R.
                Эта модель имеет следующий вид:
                Минимизировать a1*x1 + b1*x1^2 + a2*x2 + b2*x2^2
                при ограничении x1  + x2 = R
                ''')
                stringVariables = "Ваш вариант №" + str(myID) + ": "
                for a in variables:
                    stringVariables += a[0] + " = " + a[1] + "; "
                self.lVariablesText.set(stringVariables)
            else:
                if typeOfTask == '2':
                    self.lTaskText.set('''
                    Пусть p1, p2, и p3 – заданные рыночные цены трех товаров, 
                    а R – заданная константа, обозначающая личный бюджет некого субъекта. 
                    Пусть s1, s2 и s3 – определенные индивидуальные параметры, 
                    характеризующие субъекта, 
                    а x1^s1 + x2^s2 + x3^s3 обозначает “пользу”, 
                    которую субъект извлекает из потребления x1 единиц товара 1, 
                    x2 единиц товара 2 и x3 единиц товара 3. Необходимо определить 
                    потребительскую корзину, 
                    которая позволит субъекту получить максимальную пользу 
                    при соблюдении бюджетного ограничения. 
                    Математически модель можно записать в виде
                    максимизировать x1^s1 + x2^s2 + x3^s3 
                    при ограничении p1*x1 + p2*x2 + p3*x3 = R
                    ''')
                    stringVariables = "Ваш вариант № " + str(myID) + ": "
                    for a in variables:
                        stringVariables += a[0] + " = " + a[1] + "; "
                    self.lVariablesText.set(stringVariables)
                else:
                    if typeOfTask == '3':
                        self.lTaskText.set('''
                        Средние ежедневные расходы ресторана на рекламу составляют $100, 
                        причем все средства идут на рекламные объявления в газете и по радио. 
                        Обозначим через x1 среднюю сумму в день, потраченную на рекламные объявления в газете,
                        а x2 – среднюю сумму в день, потраченную на рекламу по радио. 
                        Тогда суммарные годовые затраты ресторана на содержание отдела рекламы, 
                        включая ежедневные расходы на рекламные объявления, 
                        оцениваются следующей нелинейной функцией:
                        Затраты = C(x1,x2) = K – k1*x1 – k2*x2 + k3*x1^2 + k4*x2^2 + k5*x2
                        Необходимо найти распределение бюджета ресторана, 
                        которое позволит минимизировать эти суммарные ежегодные расходы, 
                        сохранив ежедневные расходы на рекламу на уровне $100. 
                        Математическая модель имеет вид
                        минимизировать K – k1*x1 – k2*x2 + k3*x1^2 + k4*x2^2 + x1*x2
                        при ограничениях x1 + x2 = 100 и x1 ≥ 0, x2 ≥ 0.
                        ''')
                        stringVariables = "Ваш вариант № " + str(myID) + ": "
                        for a in variables:
                            stringVariables += a[0] + " = " + a[1] + "; "
                        self.lVariablesText.set(stringVariables)
                    else:
                        messagebox.showinfo('Ошибка!', 'Некорректный ввод номера задания!')
            break

        
        
    def transformToList(self, result):
        newList = []
        for b in result:
            for a in b:
                newList.append(a)
        return newList

#окно выдачи задания
class getTheory:

    numberTheory = 0

    def __init__(self,master):
        top = self.top = Toplevel(master)
        main = Frame(top,width = 600, height = 450)
        main.pack()
        
        lTitle = Label(main, text="Теория")
        lTitle.grid(row = 0, column = 5)
        
        bGet = Button(main, text = "Далее", command = self.Next)
        bGet.grid(row = 1,column = 5)
        
        self.textVariableTheory = StringVar()
        self.tTheory = Text(main, state = DISABLED, width = 60)
        self.tTheory.grid(row = 2, column = 5)

        conn = sqlite3.connect('db.db')
        self.c = conn.cursor()

        top.geometry("600x450+0+0")
        top.resizable(0,0)
    
    def Next(self):
        self.tTheory.configure(state = NORMAL)
        sql = "SELECT * FROM theory"
        result = self.c.execute(sql)
        result = self.transformToList(result)
        if self.numberTheory<len(result)-1:
            self.numberTheory += 1
        else:
            self.numberTheory = 0
        self.tTheory.delete(1.0, END)
        self.tTheory.insert(END,result[self.numberTheory])
        self.tTheory.configure(state = DISABLED)


    def transformToList(self, result):
        newList = []
        for b in result:
            for a in b:
                newList.append(a)
        return newList

#окно проверки решения
class wSolutionCheck:
    def __init__(self, master):

        self.filename = ""
        top = self.top = Toplevel(master)

        main = Frame(top,width = 400, height = 300)
        main.pack()
        lTitle = Label(main, text="Проверка решения")
        lTitle.grid(row = 0, column = 5)
        
        lBrowse = Label(main, text = "Путь к файлу: ")
        lBrowse.grid(row = 5, column = 4)

        self.ePath = Entry(main)
        self.ePath.grid(row = 5, column = 5)

        bPath = Button(main, text = "Обзор", command = self.browse)
        bPath.grid(row = 5,column = 6)

        lNumber = Label(main, text = "Номер задания: ")
        lNumber.grid(row = 6, column = 4)

        self.eNumber = Entry(main)
        self.eNumber.grid(row = 6, column = 5)

        bExCheck = Button(main, text = "Проверить", command = self.totalCheck)
        bExCheck.grid(row = 6,column = 6)

        self.lOutputValue = StringVar()
        self.lOutput = Label(main, textvariable = self.lOutputValue)
        self.lOutput.grid(row = 7, column = 5)

        top.geometry("400x300+0+0")
        top.resizable(0,0)
        conn = sqlite3.connect('db.db')
        self.c = conn.cursor()

    #проверяет расширение файла - вызывает соответствующую функцию
    def totalCheck(self):
        if self.eNumber.get().isdigit() and os.path.isfile(self.ePath.get()): 
            self.lOutputValue.set("")
            #получаем введенный путь к файлу
            filename = self.ePath.get()
            #отделяем расширение файла от пути
            filename = filename.split('.')
            #проверяем не возникло ли ошибок
            if len(filename)>2:
                messagebox.showinfo('Ошибка!', 'В пути может присутствовать только точка расширения файла!')
                print('Много точек в пути')
            else:
                #если файл имеет Excel расширение, то запускаем функцию проверки Excel файла
                if filename[len(filename)-1] == 'xlsx':
                    print('Проверка ексель файла')
                    self.excelCheck()
                else: 
                    #если файл имеет Mathcad расширение, запускаем функцию проверки Mathcad файла
                    if filename[len(filename)-1]=='xmcd':
                        print('Проверка маткад файла')
                        self.mcdxCheck()
                    else:
                        if filename[len(filename)-1] == 'py':
                            print('Проверка python файла')
                            
                        messagebox.showinfo('Ошибка!', 'Расширение файла должно быть xlsx или xmcd!')
        else: 
            messagebox.showinfo('Ошибка!', 'Проверьте номер задания и путь к файлу!')
    
    #проверяет excel файл
    def excelCheck(self):
        filename = self.ePath.get()
        task = self.eNumber.get()
        wb = load_workbook(filename)
        sheet=wb.active
        data = []
        for a in range(1,sheet.max_column+1):
            data.append([sheet.cell(1,a),sheet.cell(1,a).value,sheet.cell(2,a), sheet.cell(2,a).value])
        for a in data:
            if "целевая функция" in a:
                oFunc = a[3]
                break
        try:
            print(oFunc)
        except:
            self.lOutputValue.set(self.lOutputValue.get() + "Целевая функция не найдена.\n")
            return 0
        try:
            oFunc = re.sub('=','',oFunc)
            oFunc = re.sub(';',',',oFunc)
            oFunc = re.sub(' ','',oFunc)
            oFunc = oFunc.replace('^','**')
            oFunc = oFunc.replace('POWER','pow')
            sentence = oFunc
        except:
            print("Ошибка в обработке значения целевой функции")
        #oFunc = re.split(r'[+*/-]', oFunc)
        
        oFunc = self.findCell(oFunc)
        for a in oFunc:
            sentence = sentence.replace(a,str(self.searchValue(data, a)))
        
        sql = "SELECT * FROM tasks WHERE id LIKE "+ str(task)
        result = self.c.execute(sql)
        result = self.transformToList(result)
        oFuncDB = str(result[2]).replace('^','**')

        sql = "SELECT variables FROM tasks WHERE id LIKE "+ str(task)
        result = self.c.execute(sql)
        result = self.transformToList(result)
        result = str(result[0])
        result = result.split(';')
        for a in result:
            a = a.split(':')
            if str(self.searchValue(data,a[0]))!=a[1]:
                messagebox.showinfo("Ошибка","Проверьте введеный в таблицу " + str(a[0]))
                return 0
            else:
                print(str(self.searchValue(data,a[0]))+" = "+str(a[1]))
                oFuncDB = oFuncDB.replace(a[0],str(self.searchValue(data,a[0])))
                self.lOutputValue.set(self.lOutputValue.get() + str(a[0]) + " введен правильно. \n")
        sql = "SELECT * FROM tasks WHERE id LIKE "+ str(task)
        result = self.c.execute(sql)
        result = self.transformToList(result)
        oFunc = str(result[2])
        
        x1 = x = self.searchValue(data,'x1')
        x2 = y = self.searchValue(data,'x2')
        x3 = z = self.searchValue(data,'x3')
        if eval(sentence) == eval(oFuncDB):
            self.lOutputValue.set(self.lOutputValue.get() + " Целевая функция правильная.\n ")
            print("Целевая функция правильная")
        if self.limitCheck(data)==1:
            self.lOutputValue.set(self.lOutputValue.get() + " Ограничение верно!\n ")
            print("Ограничение найдено")
        else:
            self.lOutputValue.set(self.lOutputValue.get() + " Ограничение с ошибкой!\n ")
            print("Ошибка с ограничением")
        

    def transformToList(self, result):
        newList = []
        for b in result:
            for a in b:
                newList.append(a)
        return newList

    #Ищет target по двумерному массиву data и возвращает значение
    def searchValue(self,data, target):
        for a in data:
            if re.search(target,str(a[2]))!=None:
                    return a[3]
            if re.search(target,str(a[1]))!=None:
                    return a[3]

    #обзор файла
    def browse(self):
        Tk().withdraw() 
        self.filename = filedialog.askopenfilename()
        self.ePath.delete(0,'end')
        self.ePath.insert(0,self.filename)

    def findCell(self, sentence):
        alphabit = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
        result = []
        for symbol in alphabit:
            for a in range(0,10):
                if (symbol + str(a)) in sentence:
                    result.append(symbol + str(a))
        return result
    
    #вырезает из data кусок, начинающийся на start и заканчивающийся на end. Возвращает список вырезанных кусков
    #ножницы
    def scissors(self,data,start,end):
        result = []
        #пока в тексте для поиска существуют последовательности start и end, цикл продолжается
        while data.find(start)!=-1 and data.find(end)!=-1:
            #в список результата добавляется найденная конструкция
            result.append(data[data.index(start):data.index(end)+len(end)])
            #найденная конструкция вырезается из текста для поиска
            data = data[data.index(end)+len(end):]
        #возвращаем список найденных конструкций
        return result
        
    #проверка файла маткада
    def mcdxCheck(self):
        filename = self.ePath.get()
        task = self.eNumber.get()
        with open(filename,'r') as content_file:
            data2 = content_file.read()
            data2 = data2.replace(" xml:space=\"preserve\"", "").replace(" style=\"auto-select\"","").replace('\t',"").replace(" xmlns:ml=\"http://schemas.mathsoft.com/math30\"","").replace(" warning=\"WarnRedefinedBIUnit\"","")
            data = self.scissors(data2, '<ml:define','</ml:define>')
            sql = "SELECT variables FROM tasks WHERE id LIKE "+ str(task)
            result = self.c.execute(sql)
            result = self.transformToList(result)[0].split(';')
            variables = []
            varMcd = []
            sql = "SELECT [typeOfTask] FROM tasks WHERE id LIKE "+ str(task)
            types = self.c.execute(sql)
            typeOT =[a[0] for a in types][0]
            print("Проверка маткад файла")
            if "<ml:id>x1</ml:id>\n<ml:real>" in  data2: 
                self.lOutputValue.set(self.lOutputValue.get() + "x1 на месте!\n")
                if "<ml:id>x2</ml:id>\n<ml:real>" in  data2: 
                    self.lOutputValue.set(self.lOutputValue.get() + "x2 на месте!\n")
                    if "<ml:id>x3</ml:id>\n<ml:real>" in  data2 and  typeOT == '2': 
                        self.lOutputValue.set(self.lOutputValue.get() + "x3 на месте!\n")
                    else:
                        if typeOT == "2":
                            self.lOutputValue.set(self.lOutputValue.get() + "Задайте начальное значение x3!\n")
                            return 0
                else:
                    self.lOutputValue.set(self.lOutputValue.get() + "Задайте начальное значение x2!\n")
                    return 0
            else:
                self.lOutputValue.set(self.lOutputValue.get() + "Задайте начальное значение x1!\n")
                return 0


            for a in result:
                variables.append(a.split(":"))
            for a in variables:
                varMcd.append(self.transformVariableToMcd(a))
            for a in varMcd:
                if a in data:
                    self.lOutputValue.set(self.lOutputValue.get() + a[a.index("<ml:id>")+len("<ml:id>"):a.index("</ml:id>")] + " на месте!\n")
                    data.remove(a)
                else:
                    print(a[a.index("<ml:id>")+len("<ml:id>"):a.index("</ml:id>")] + " отсутствует!")
                    self.lOutputValue.set(self.lOutputValue.get() + a[a.index("<ml:id>")+len("<ml:id>"):a.index("</ml:id>")] + " отсутствует!\n")
                    return 0
            sql = "SELECT function FROM tasks WHERE id LIKE "+ str(task)
            result = self.c.execute(sql)
            result = self.transformToList(result)[0].split(';')
            ofMcd = self.transformOFtoMcd(result[0])
            
            if ofMcd in data:  
                print("Целевая функция найдена!")
                self.lOutputValue.set(self.lOutputValue.get() + " Целевая функция задана правильно.\n ")
                #data.remove(a)
            else:
                print("Целевая функция с ошибкой!")
                self.lOutputValue.set(self.lOutputValue.get() + " Целевая функция задана неправильно.\n ")
                return 0
            limit = self.transformLimitToMcd()

            if "<ml:id>Given</ml:id>\n" in data2:
                print("Given на месте")
                self.lOutputValue.set(self.lOutputValue.get() + " Given присутствует.\n ")
            else:
                print("Given не найдено!")
                self.lOutputValue.set(self.lOutputValue.get() + " Given отсутствует.\n ")
                return 0

            if limit in data2:
                print("Оганичение найдено!")
                self.lOutputValue.set(self.lOutputValue.get() + " Ограничение задано правильно.\n ")
            else:
                print("Оганичение не найдено!")
                self.lOutputValue.set(self.lOutputValue.get() + " Ограничение задано неправильно.\n ")
                return 0
                

            
    #проверка ограничения в Excel
    def limitCheck(self, book):
        task = self.eNumber.get()
        sql = "SELECT [limit] FROM tasks WHERE id LIKE "+ str(task)
        limit = self.c.execute(sql)
        limit = self.transformToList(limit)
        limit = str(limit[0])
        if '<=' in limit:
            limit = limit.split('<=')
            znak = '<='
        else:
            if '>=' in limit:
                limit = limit.split('>=')
                znak = '>='
            else:
                if '=' in limit:
                    limit = limit.split('=')
                    znak = '='
                else:
                    if '<' in limit:
                        limit = limit.split('<')
                        znak = '<'
                    else:
                        if '>' in limit:
                            limit = limit.split('>')
                            znak = '>'
        sql = "SELECT variables FROM tasks WHERE id LIKE "+ str(task)
        result = self.c.execute(sql)
        result = self.transformToList(result)[0].split(';')
        variables = []
        for a in result:
            variables.append(a.split(":"))
        for a in variables:
            if a[0] in limit[0]:
                limit[0] = limit[0].replace(a[0],a[1])
        for a in variables:
            for b in a:
                if limit[1] in b:
                    index = variables.index(a)
                    break
        for a in book:
            if str(a[1]) in limit[0]:
                limit[0] = limit[0].replace(a[1],str(a[3]))
        if znak == '<=':
            if eval(limit[0]) <= int(variables[index][1]):
                print("Ограничение проверено!")
                return 1
        else:
            if znak == '>=':
                if eval(limit[0]) >= int(variables[index][1]):
                    print("Ограничение проверено!")
                    return 1
            else:
                if znak == '=':
                    if eval(limit[0]) == int(variables[index][1]):
                        print("Ограничение проверено!")
                        return 1
                else:
                    if znak == '<':
                        if eval(limit[0]) < int(variables[index][1]):
                            print("Ограничение проверено!")
                            return 1
                    else:
                        if znak == '>':
                            if eval(limit[0]) > int(variables[index][1]):
                                print("Ограничение проверено!")
                                return 1
                        else:
                            print("Ограничение неверно!")
                            return 0
    
    #трансформирует инициализацию переменной из листа в маткадовский синтаксис
    def transformVariableToMcd(self, variable):
        return "<ml:define>\n<ml:id>"+variable[0]+"</ml:id>\n<ml:real>"+variable[1]+"</ml:real>\n</ml:define>"

    #трансформирует целевую функцию из бд в синтаксис маткада
    def transformOFtoMcd(self,oFunc):
        task = self.eNumber.get()
        oFunc = oFunc.replace(' ','').replace('/','|')
        sql = "SELECT [typeOfTask] FROM tasks WHERE id LIKE "+ str(task)
        types = self.c.execute(sql)
        typeOT =[a[0] for a in types][0]
        if typeOT == 2:
            forType2 = "<ml:id>x3</ml:id>\n"
        else:
            forType2 = ""
        result = "<ml:define>\n<ml:function>\n<ml:id>f</ml:id>\n<ml:boundVars>\n<ml:id>x1</ml:id>\n<ml:id>x2</ml:id>\n"+forType2+"</ml:boundVars>\n</ml:function>\n"
        '''
        quote = 0
        for a in oFunc:
            if a == '(' or a == ')':
                quote+=1
        while oFunc.find('(')!=-1:
            part = oFunc[oFunc.index('(')+1:oFunc.index(')')]
            save = part
            if part.find('+')!=-1: 
                part = part.split('+')
                part.append('+')
            else:
                if part.find('-')!=-1: 
                    part = part.split('-')
                    part.append('-')
                else:
                    if part.find('**')!=-1:
                        part = part.split('**')
                        part.append('**')
                    else:
                        if part.find('|')!=-1:
                            part = part.split('|')
                            part.append('|')
                        else:
                            if part.find('*')!=-1:
                                part = part.split('*')
                                part.append('*')
                            else:
                                if part.find('^')!=-1:
                                    part = part.split('^')
                                    part.append('^')
            sentence = ""
            if part[2]=='+':
                sentence += '<ml:apply>\n<ml:plus/>\n'
            else:
                if part[2]=='-':
                    sentence += '<ml:apply>\n<ml:minus/>\n'
                else:
                    if part[2] == '*':
                        sentence += '<ml:apply>\n<ml:mult/>\n'
                    else:
                        if part[2] == '|':
                            sentence+='<ml:apply>\n<ml:div/>\n'
                        else:
                            if part[2] == '**' or part[2] == '^':
                                sentence+='<ml:apply>\n<ml:pow/>\n'
            sentence += self.typeCheck(part[0])
            sentence += self.typeCheck(part[1])
            sentence += "</ml:apply>\n"
            oFunc = oFunc.replace('('+save+')', sentence)
            '''
        
        masOfunc = oFunc.replace('^','?^?').replace('**','?^?').replace('+','?+?').replace('-','?-?').replace('*','?*?').replace('|','?|?').split('?')
        oFunc = oFunc.replace('**','^')

        signs = ["^","**","*","|","+","-"]

        while '**' in masOfunc or '^' in masOfunc:
            index = 0
            for a in masOfunc:
                if masOfunc[index]=='**' or masOfunc[index]=='^':
                    signS = "<ml:pow/>\n"
                    sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                    oFunc = oFunc.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                    masOfunc.pop(index-1)
                    masOfunc.pop(index-1)
                    masOfunc[index-1] = sentence
                    index = 0
                index+=1
        index = 1
        while '*' in masOfunc or '|' in masOfunc:
            index = 0
            for a in masOfunc:
                if masOfunc[index]=='*':
                    signS = "<ml:mult/>\n"
                    sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                    oFunc = oFunc.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                    masOfunc.pop(index-1)
                    masOfunc.pop(index-1)
                    masOfunc[index-1] = sentence
                    index = 0
                else:
                    if masOfunc[index]=='|':
                        signS = "<ml:div/>\n"
                        sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                        oFunc = oFunc.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                        masOfunc.pop(index-1)
                        masOfunc.pop(index-1)
                        masOfunc[index-1] = sentence
                        index = 0
                index+=1
        
        while '+' in masOfunc or '-' in masOfunc:
            index = 0
            for a in masOfunc:
                if masOfunc[index]=='+':
                    signS = "<ml:plus/>\n"
                    sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                    oFunc = oFunc.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                    masOfunc.pop(index-1)
                    masOfunc.pop(index-1)
                    masOfunc[index-1] = sentence
                    index = 0
                else:
                    if masOfunc[index]=='-':
                        signS = "<ml:minus/>\n"
                        sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                        oFunc = oFunc.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                        masOfunc.pop(index-1)
                        masOfunc.pop(index-1)
                        masOfunc[index-1] = sentence
                        index = 0
                index+=1
        return result+oFunc.replace('!','-')+'</ml:define>'

    #трансформирует ограничение из бд в синтаксис маткада
    def transformLimitToMcd(self):
        task = self.eNumber.get()
        sql = "SELECT [limit] FROM tasks WHERE id LIKE "+ str(task)
        limit = self.c.execute(sql)
        limit = self.transformToList(limit)
        limit = str(limit[0])
        if '<=' in limit:
            limit = limit.split('<=')
            znak = '<ml:lessOrEqual/>\n'
        else:
            if '>=' in limit:
                limit = limit.split('>=')
                znak = '<ml:greaterOrEqual/>\n'
            else:
                if '=' in limit:
                    limit = limit.split('=')
                    znak = '<ml:equal/>\n'
                else:
                    if '<' in limit:
                        limit = limit.split('<')
                        znak = '<ml:lessThan/>\n'
                    else:
                        if '>' in limit:
                            limit = limit.split('>')
                            znak = '<ml:greaterThan/>\n'
        result = '<ml:apply>\n' + znak
        R = limit[1]
        limit = limit[0]
        while limit.find('(')!=-1:
            part = limit[len(limit)-limit.rindex('(')+2:limit.index(')')]
            save = part
            if part.find('+')!=-1: 
                part = part.split('+').append('+')
            else:
                if part.find('-')!=-1: 
                    part = part.split('-').append('-')
                else:
                    if part.find('**')!=-1:
                        part = part.split('**').append('**')
                    else:
                        if part.find('|')!=-1:
                            part = part.split('|').append('|')
                        else:
                            if part.find('*')!=-1:
                                part = part.split('*').append('*')
                            else:
                                if part.find('^')!=-1:
                                    part = part.split('^').append('^')
            sentence = ""
            if part[2]=='+':
                sentence += '<ml:apply>\n<ml:plus/>\n'
            else:
                if part[2]=='-':
                    sentence += '<ml:apply>\n<ml:minus/>\n'
                else:
                    if part[2] == '*':
                        sentence += '<ml:apply>\n<ml:mult/>\n'
                    else:
                        if part[2] == '|':
                            sentence+='<ml:apply>\n<ml:div/>\n'
                        else:
                            if part[2] == '**' or part[2] == '^':
                                sentence+='<ml:apply>\n<ml:pow/>\n'
            sentence += self.typeCheck(part[0])
            sentence += self.typeCheck(part[1])
            sentence += "</ml:apply>\n"
            limit = limit.replace(save, sentence)
        
        masOfunc = limit.replace('^','?^?').replace('**','?^?').replace('+','?+?').replace('-','?-?').replace('*','?*?').replace('|','?|?').split('?')
        signs = ["^","**","*","|","+","-"]

        while '**' in masOfunc or '^' in masOfunc:
            index = 0
            for a in masOfunc:
                if masOfunc[index]=='**' or masOfunc[index]=='^':
                    signS = "<ml:pow/>\n"
                    sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                    limit = limit.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                    masOfunc.remove(masOfunc[index-1])
                    masOfunc.remove(masOfunc[index-1])
                    masOfunc[index-1] = sentence
                index+=1
        index = 1
        while '*' in masOfunc or '|' in masOfunc:
            index = 0
            for a in masOfunc:
                if masOfunc[index]=='*':
                    signS = "<ml:mult/>\n"
                    sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                    limit = limit.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                    masOfunc.remove(masOfunc[index-1])
                    masOfunc.remove(masOfunc[index-1])
                    masOfunc[index-1] = sentence
                else:
                    if masOfunc[index]=='|':
                        signS = "<ml:div/>\n"
                        sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                        limit = limit.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                        masOfunc.remove(masOfunc[index-1])
                        masOfunc.remove(masOfunc[index-1])
                        masOfunc[index-1] = sentence
                index+=1
        
        while '+' in masOfunc or '-' in masOfunc:
            index = 0
            for a in masOfunc:
                if masOfunc[index]=='+':
                    signS = "<ml:plus/>\n"
                    sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                    limit = limit.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                    masOfunc.remove(masOfunc[index-1])
                    masOfunc.remove(masOfunc[index-1])
                    masOfunc[index-1] = sentence
                else:
                    if masOfunc[index]=='-':
                        signS = "<ml:minus/>\n"
                        sentence = "<ml:apply>\n"+signS+self.typeCheck(masOfunc[index-1])+self.typeCheck(masOfunc[index+1])+"</ml:apply>\n"
                        limit = limit.replace(masOfunc[index-1]+masOfunc[index]+masOfunc[index+1],sentence)
                        masOfunc.remove(masOfunc[index-1])
                        masOfunc.remove(masOfunc[index-1])
                        masOfunc[index-1] = sentence
                index+=1
        return result+limit.replace('!','-')+self.typeCheck(R)+'</ml:apply>\n'

    #определяет является ли data числом, переменной или нет. возвращает обернутые в теги
    def typeCheck(self,data):
        #получаем вариант
        task = self.eNumber.get()
        #собираем SQL-запрос для базы данных
        sql = "SELECT variables FROM tasks WHERE id LIKE " + str(task)
        #выполняем собранную команду, получаем ответ
        result = self.c.execute(sql)
        #преобразовываем полученный ответ
        result = self.transformToList(result)[0].split(';')
        variables = []
        #еще раз преобразовываем
        for a in result:
            variables.append(a.split(":"))
        #составляем регулярное выражение, которое проверяет строку на число
        p = re.compile('^-?[0-9]\d*(\.\d+)?$')
        #проверяем, можно ли преобразовать строку в число
        if p.match(data)!=None:
            #возвращаем данные, обернутые в теги для числа
            return "<ml:real>"+ str(data)+"</ml:real>\n"
        else:
            #проверяем является ли данная строка переменной из задания
            for a in variables:
                if a[0] == data:
                    #возвращаем данные, обернутые в теги для переменной
                    return "<ml:id>"+str(data)+"</ml:id>\n"
            #проверяем является ли данная строка переменной из задания
            if data == 'x' or data == 'y' or data == 'z'or data == 'x1'or data == 'x2'or data == 'x3':
                #возвращаем данные, обернутые в теги для переменной
                return "<ml:id>"+str(data)+"</ml:id>\n"
            else:
                #возвращаем результат без изменений
                return data

root = Tk()
b = Application(root)
root.geometry("800x600+0+0")
root.resizable(0,0)
root.mainloop()
